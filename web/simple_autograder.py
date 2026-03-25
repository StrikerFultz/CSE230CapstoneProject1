#!/usr/bin/env python3

import psycopg2
from psycopg2.extras import RealDictCursor
from flask import Blueprint, request, jsonify
from flask import session

import json
import subprocess
import tempfile
import os
import platform

simple_autograder_bp = Blueprint('simple_autograder', __name__, url_prefix='/api/grade')

_binary_name = 'mips-emu-wasm.exe' if platform.system() == 'Windows' else 'mips-emu-wasm'
GRADER_BINARY = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', 'target', 'release',
    _binary_name
)

DB_CONFIG = {
    'dbname': 'capstone',
    'user': 'postgres',
    'password': 'postgres',
    'host': 'localhost',
    'port': '5432'
}

def run_mips_native(source_code, initial_registers=None, initial_memory=None, check_memory=None):
    """
    Run student code using the compiled emulator binary.
    """

    if initial_registers is None:
        initial_registers = {}
    if initial_memory is None:
        initial_memory = {}
    if check_memory is None:
        check_memory = []

    payload = json.dumps({
        'source_code': source_code,
        'initial_registers': initial_registers,
        'initial_memory': initial_memory,
        'check_memory': check_memory,
    })

    try:
        result = subprocess.run(
            [GRADER_BINARY],
            input=payload,
            capture_output=True,
            text=True,
            timeout=10
        )

        if result.returncode != 0:
            stderr_msg = result.stderr.strip() if result.stderr else 'Unknown error'
            if result.stdout.strip():
                try:
                    return json.loads(result.stdout)
                except json.JSONDecodeError:
                    pass
            return {'error': f'Binary exited with code {result.returncode}: {stderr_msg}', 'registers': {}, 'memory': {}}

        output = json.loads(result.stdout)
        return output

    except Exception as e:
        print(f"[GRADER] Error: {e}")
        return {'error': str(e), 'registers': {}, 'memory': {}}

def calculate_grade(test_cases, source_code):
    """Grade by running source code against each test case via the Rust binary."""

    total_points = 0
    earned_points = 0
    passed = 0
    failed = 0

    results = []
    
    for test in test_cases:
        test_name = test.get('name', 'Test')
        points = test.get('points', 10)

        initial_regs = test.get('initial_registers', {})
        initial_mem = test.get('initial_memory', {})

        expected_regs = test.get('expected_registers', {})
        expected_mem = test.get('expected_memory', {})
        
        total_points += points

        # Build list of memory addresses we need to check
        check_memory = [int(addr) for addr in expected_mem.keys()] if expected_mem else []

        # Run the student's code with initial values using the emulator binary
        run_result = run_mips_native(
            source_code,
            initial_registers=initial_regs,
            initial_memory=initial_mem,
            check_memory=check_memory
        )

        if run_result.get('error'):
            failed += 1
            results.append({
                'name': test_name,
                'status': 'ERROR',
                'points': points,
                'earned': 0,
                'message': f'✗ Runtime error: {run_result["error"]}'
            })
            continue
        
        student_registers = run_result.get('registers', {})
        
        # Check if results match expected
        all_correct = True
        mismatches = []
        
        for register, expected_value in expected_regs.items():
            student_value = student_registers.get(register)
            
            if student_value != expected_value:
                all_correct = False
                mismatches.append({
                    'register': register,
                    'expected': expected_value,
                    'actual': student_value
                })

        
        student_memory = run_result.get('memory', {})

        for addr_str, expected_value in expected_mem.items():
            student_value = student_memory.get(str(addr_str))

            if student_value != expected_value:
                all_correct = False
                mismatches.append({
                    'register': f'mem[{addr_str}]',
                    'expected': expected_value,
                    'actual': student_value
                })

        
        if all_correct:
            earned_points += points
            passed += 1
            results.append({
                'name': test_name,
                'status': 'PASS',
                'points': points,
                'earned': points,
                'message': 'All checks passed!'
            })
        else:
            failed += 1
            results.append({
                'name': test_name,
                'status': 'FAIL',
                'points': points,
                'earned': 0,
                'message': 'Some values incorrect',
                'mismatches': mismatches
            })
    
    percentage = (earned_points / total_points * 100) if total_points > 0 else 0
    
    return {
        'earned_points': earned_points,
        'total_points': total_points,
        'percentage': round(percentage, 1),
        'passed': passed,
        'failed': failed,
        'results': results
    }

def get_test_cases_for_lab(lab_id):
    """
    Get test cases from database and convert to the flat format
    that calculate_grade() expects.
    """
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT test_name, description, input_data, expected_output, points, is_hidden
            FROM lab_test_cases
            WHERE lab_id = %s
            ORDER BY test_name
        """, (lab_id,))
        
        test_cases = cursor.fetchall()
        cursor.close()
        conn.close()

        if not test_cases:
            return None
        
        # Convert DB JSONB → flat format for the grader
        result = []
        for tc in test_cases:
            # Parse JSONB (psycopg2 usually returns dicts, but guard against strings)
            inp = tc.get('input_data') or {}
            exp = tc.get('expected_output') or {}
            if isinstance(inp, str):
                inp = json.loads(inp)
            if isinstance(exp, str):
                exp = json.loads(exp)

            result.append({
                'name':               tc['test_name'],
                'description':        tc.get('description'),
                'points':             tc.get('points', 10),
                'initial_registers':  inp.get('registers', {}),
                'initial_memory':     inp.get('memory', {}),
                'expected_registers': exp.get('registers', {}),
                'expected_memory':    exp.get('memory', {}),
                'is_hidden':          tc.get('is_hidden', False),
            })
        
        return result
        
    except Exception as e:
        print(f"[AUTOGRADER] Database error fetching test cases for {lab_id}: {e}")
        return None


@simple_autograder_bp.route('/submit', methods=['POST'])
def grade_submission():
    """
    POST /api/grade/submit
    
    Body: #how setup woroks
    {
        "lab_id": "lab-12-2", # get lab type (named in db)
        "source_code": "add $t0, $s0, $s1\\n..." 
    }
    """
        
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json() 
        lab_id = data.get('lab_id') 
        source_code = data.get('source_code', '') 
        
        test_cases = get_test_cases_for_lab(lab_id) 
        grade_report = calculate_grade(test_cases, source_code) 

        # Save to DB
        conn = psycopg2.connect(**DB_CONFIG) 
        cur = conn.cursor() #

        normalized_source = source_code.strip().replace('\n', '\\n')

        cur.execute("""
            INSERT INTO submissions (user_id, asurite_id, lab_id, score, total_possible, source_code, test_results)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            session['user_id'],
            session.get('username'), # ASU ID
            lab_id,
            grade_report['earned_points'],
            grade_report['total_points'],
            normalized_source, 
            json.dumps(grade_report['results'])
        ))

        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'lab_id': lab_id,
            'grade_report': grade_report
        })  
          
    except Exception as e:
        return jsonify({'error': str(e)}), 500 #

@simple_autograder_bp.route('/test-cases/<lab_id>', methods=['GET'])
def get_test_cases_endpoint(lab_id):
    """Returns test cases WITHOUT expected answers (safe for students)."""
    test_cases = get_test_cases_for_lab(lab_id)

    if not test_cases:
        return jsonify({'error': f'No test cases found for {lab_id}'}), 404
    
    # Remove expected answers before sending to student
    sanitized = []
    for test in test_cases:
        sanitized.append({
            'name': test.get('name'),
            'description': test.get('description'),
            'points': test.get('points')
        })
    
    return jsonify({
        'lab_id': lab_id,
        'test_cases': sanitized
    })


@simple_autograder_bp.route('/latest/<lab_id>', methods=['GET'])
def get_latest_submission(lab_id):
    """Fetches the most recent attempt for the logged-in student."""
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT score, total_possible, test_results, submitted_at, source_code
        FROM submissions
        WHERE user_id = %s AND lab_id = %s
        ORDER BY submitted_at DESC LIMIT 1
    """, (session['user_id'], lab_id))
    
    row = cur.fetchone()
    cur.close()
    conn.close()
    
    if not row:
        return jsonify({'error': 'No submissions found'}), 404
        
    return jsonify({
        'score': row['score'],
        'total_possible': row['total_possible'],
        'test_results': row['test_results'],
        'submitted_at': row['submitted_at'].isoformat(),
        'source_code': row['source_code']
    })

#^ autograder from wasm from march 5(?)
# new code for downloading graders(csv) and history of student submission

from io import BytesIO, StringIO
from flask import send_file
import zipfile
import csv
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARNING] openpyxl not installed. Excel exports will not work. Install with: pip3 install openpyxl")


@simple_autograder_bp.route('/export/grades/<lab_id>', methods=['GET'])
def export_grades_excel(lab_id):
    """Download best grades as Excel with manual adjustment column"""
    
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel library not installed'}), 500
    
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get best submission for each student
        if lab_id == "ALL":
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id, s.lab_id)
                    u.username,
                    u.user_id,
                    u.email,
                    s.lab_id,
                    s.score,
                    s.total_possible as max_score,
                    (s.score::float / s.total_possible * 100) as percentage,
                    s.submitted_at
                FROM submissions s
                JOIN users u ON s.user_id = u.user_id
                ORDER BY s.user_id, s.lab_id, s.score DESC, s.submitted_at DESC
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id)
                    u.username,
                    u.user_id,
                    u.email,
                    s.score,
                    s.total_possible as max_score,
                    (s.score::float / s.total_possible * 100) as percentage,
                    s.submitted_at
                FROM submissions s
                JOIN users u ON s.user_id = u.user_id
                WHERE s.lab_id = %s
                ORDER BY s.user_id, s.score DESC, s.submitted_at DESC
            """, (lab_id,))
        
        submissions = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not submissions:
            return jsonify({'error': 'No submissions found for this lab'}), 404
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        
        # Headers
        if lab_id == "ALL":
            headers = ['Student Name', 'Student ID', 'Email', 'Lab', 'Auto Score', 
                      'Max Score', 'Auto %', 'Manual Adj', 'Final Score', 'Final %', 'Submitted']
        else:
            headers = ['Student Name', 'Student ID', 'Email', 'Auto Score', 
                      'Max Score', 'Auto %', 'Manual Adj', 'Final Score', 'Final %', 'Submitted']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, sub in enumerate(submissions, 2):
            col = 1
            ws.cell(row=row_idx, column=col, value=sub.get('username', 'Unknown'))
            col += 1
            ws.cell(row=row_idx, column=col, value=str(sub['user_id']))
            col += 1
            ws.cell(row=row_idx, column=col, value=sub.get('email', ''))
            col += 1
            
            if lab_id == "ALL":
                ws.cell(row=row_idx, column=col, value=sub.get('lab_id', ''))
                col += 1
            
            # Auto Score
            auto_score_col = col
            ws.cell(row=row_idx, column=col, value=sub['score'])
            col += 1
            
            # Max Score
            max_score_col = col
            ws.cell(row=row_idx, column=col, value=sub['max_score'])
            col += 1
            
            # Auto %
            pct_cell = ws.cell(row=row_idx, column=col, value=float(sub.get('percentage', 0)))
            if sub.get('percentage', 0) >= 90:
                pct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif sub.get('percentage', 0) >= 70:
                pct_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                pct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1
            
            # Manual Adjustment (YELLOW - editable)
            adj_cell = ws.cell(row=row_idx, column=col, value=0.0)
            adj_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            adj_cell.font = Font(bold=True)
            adj_col = col
            col += 1
            
            # Final Score (Formula)
            final_cell = ws.cell(row=row_idx, column=col)
            final_cell.value = f"={chr(64 + auto_score_col)}{row_idx}+{chr(64 + adj_col)}{row_idx}"
            final_cell.font = Font(bold=True)
            final_col = col
            col += 1
            
            # Final % (Formula)
            final_pct_cell = ws.cell(row=row_idx, column=col)
            final_pct_cell.value = f"=({chr(64 + final_col)}{row_idx}/{chr(64 + max_score_col)}{row_idx})*100"
            final_pct_cell.number_format = '0.0'
            final_pct_cell.font = Font(bold=True)
            col += 1
            
            # Submitted At
            ws.cell(row=row_idx, column=col, value=sub['submitted_at'].strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'grades_{lab_id}_{timestamp}.xlsx' if lab_id != "ALL" else f'all_grades_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting grades: {e}")
        return jsonify({'error': str(e)}), 500


@simple_autograder_bp.route('/export/submissions-zip/<lab_id>', methods=['GET'])
def export_submissions_zip(lab_id):
    """Download all student source code as ZIP"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        if lab_id == "ALL":
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id, s.lab_id)
                    u.username,
                    u.user_id,
                    s.lab_id,
                    s.source_code,
                    s.score,
                    s.submitted_at
                FROM submissions s
                JOIN users u ON s.user_id = u.user_id
                ORDER BY s.user_id, s.lab_id, s.score DESC
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id)
                    u.username,
                    u.user_id,
                    s.lab_id,
                    s.source_code,
                    s.score,
                    s.submitted_at
                FROM submissions s
                JOIN users u ON s.user_id = u.user_id
                WHERE s.lab_id = %s
                ORDER BY s.user_id, s.score DESC
            """, (lab_id,))
        
        submissions = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not submissions:
            return jsonify({'error': 'No submissions found'}), 404
        
        # Create ZIP
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for sub in submissions:
                filename = f"{sub['username']}_{sub['lab_id']}_score{sub['score']}.asm"
                file_content = f"""# Student: {sub['username']}
# Student ID: {sub['user_id']}
# Lab: {sub['lab_id']}
# Score: {sub['score']}
# Submitted: {sub['submitted_at']}
# ==========================================

{sub['source_code']}
"""
                zip_file.writestr(filename, file_content)
        
        zip_buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'submissions_{lab_id}_{timestamp}.zip' if lab_id != "ALL" else f'all_submissions_{timestamp}.zip'
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error creating ZIP: {e}")
        return jsonify({'error': str(e)}), 500


@simple_autograder_bp.route('/export/submissions/<lab_id>', methods=['GET'])
def export_submissions_csv(lab_id):
    """Download complete submission history as CSV"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        if lab_id == "ALL":
            cursor.execute("""
                SELECT 
                    u.username,
                    u.user_id,
                    u.email,
                    s.lab_id,
                    s.score,
                    s.total_possible,
                    s.source_code,
                    s.submitted_at
                FROM submissions s
                JOIN users u ON s.user_id = u.user_id
                ORDER BY u.username, s.lab_id, s.submitted_at DESC
            """)
        else:
            cursor.execute("""
                SELECT 
                    u.username,
                    u.user_id,
                    u.email,
                    s.score,
                    s.total_possible,
                    s.source_code,
                    s.submitted_at
                FROM submissions s
                JOIN users u ON s.user_id = u.user_id
                WHERE s.lab_id = %s
                ORDER BY u.username, s.submitted_at DESC
            """, (lab_id,))
        
        submissions = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not submissions:
            return jsonify({'error': 'No submissions found'}), 404
        
        # Create CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        if lab_id == "ALL":
            writer.writerow(['Student Name', 'Student ID', 'Email', 'Lab', 'Score', 
                           'Max Score', 'Percentage', 'Submitted At', 'Source Code'])
        else:
            writer.writerow(['Student Name', 'Student ID', 'Email', 'Score', 
                           'Max Score', 'Percentage', 'Submitted At', 'Source Code'])
        
        # Data
        for sub in submissions:
            percentage = (sub['score'] / sub['total_possible'] * 100) if sub['total_possible'] > 0 else 0
            row = [
                sub.get('username', 'Unknown'),
                str(sub['user_id']),
                sub.get('email', ''),
            ]
            
            if lab_id == "ALL":
                row.append(sub.get('lab_id', ''))
            
            row.extend([
                sub['score'],
                sub['total_possible'],
                f"{percentage:.1f}%",
                sub['submitted_at'].strftime('%Y-%m-%d %H:%M:%S'),
                sub['source_code'].replace('\n', ' | ')
            ])
            
            writer.writerow(row)
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'history_{lab_id}_{timestamp}.csv' if lab_id != "ALL" else f'all_history_{timestamp}.csv'
        
        return send_file(
            BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting CSV: {e}")
        return jsonify({'error': str(e)}), 500
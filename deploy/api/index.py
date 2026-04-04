import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import os
import sys
import json
import csv
import logging
import zipfile
from io import BytesIO, StringIO
from datetime import datetime

from dotenv import load_dotenv
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    stream=sys.stderr,
)

from flask import Flask, jsonify, request, session, send_file
from flask_cors import CORS
from psycopg2.extras import RealDictCursor

from _db import get_db_connection
from _auth import auth_bp
from _autograder import simple_autograder_bp

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)

app.register_blueprint(auth_bp)
app.register_blueprint(simple_autograder_bp)

_secret = os.environ.get('FLASK_SECRET_KEY')
if not _secret:
    if os.environ.get('VERCEL_ENV') == 'production':
        raise RuntimeError('FLASK_SECRET_KEY must be set in production.')
    _secret = 'dev-secret-key-change-me'

app.secret_key = _secret

app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE']   = os.environ.get('VERCEL_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY']  = True

CORS(app, supports_credentials=True)


def _require_login():
    uid = session.get('user_id')
    if not uid:
        return None, None
    return uid, session.get('role', 'student')

def _require_teacher():
    uid, role = _require_login()
    if not uid:
        return None, 'not_logged_in'
    if role not in ('instructor', 'ta'):
        return None, 'not_teacher'
    return uid, None

def _safe_error(e, fallback='An internal error occurred'):
    app.logger.error('%s', e, exc_info=True)
    return fallback

def _fetch_test_cases(cursor, lab_id):
    cursor.execute("""
        SELECT test_case_id, test_name, test_type, description,
               input_data, expected_output, points, is_hidden, timeout_seconds
        FROM lab_test_cases
        WHERE lab_id = %s ORDER BY test_name
    """, (lab_id,))
    rows = cursor.fetchall()

    test_cases = []
    for row in rows:
        inp = row.get('input_data') or {}
        exp = row.get('expected_output') or {}
        if isinstance(inp, str): inp = json.loads(inp)
        if isinstance(exp, str): exp = json.loads(exp)

        test_cases.append({
            'test_case_id':    str(row['test_case_id']),
            'test_name':       row['test_name'],
            'test_type':       row.get('test_type', 'register'),
            'description':     row.get('description'),
            'input_data':      inp,
            'expected_output': exp,
            'points':          row.get('points', 10),
            'is_hidden':       row.get('is_hidden', False),
            'timeout_seconds': row.get('timeout_seconds', 5),
        })
    return test_cases


@app.route('/api/test-connection', methods=['GET'])
def test_connection():
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT version();")
            version = cursor.fetchone()
            conn.close()
            return jsonify({'status': 'success', 'message': 'Database connected',
                            'version': version[0]})
        except Exception as e:
            return jsonify({'status': 'error', 'message': _safe_error(e)}), 500
    return jsonify({'status': 'error', 'message': 'Could not connect'}), 500


@app.route('/api/labs', methods=['GET'])
def get_labs():
    uid, role = _require_login()
    if not uid:
        return jsonify({'error': 'Not logged in'}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    is_teacher = role in ('instructor', 'ta')

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT lab_id, course_id, title, description, instructions,
                starter_code, register_mapping, initial_values,
                max_memory_kb, time_limit_seconds, max_instructions,
                total_points, release_date, due_date, is_published, difficulty,
                use_isolation
            FROM labs ORDER BY lab_id
        """)
        labs = cursor.fetchall()

        result = {}
        for lab in labs:
            lab_id = lab['lab_id']
            test_cases = _fetch_test_cases(cursor, lab_id) if is_teacher else []

            reg_map   = lab.get('register_mapping') or {}
            init_vals = lab.get('initial_values') or {}

            result[lab_id] = {
                'title':            lab['title'],
                'html':             lab.get('instructions') or '',
                'description':      lab.get('description'),
                'starter_code':     lab.get('starter_code'),
                'difficulty':       lab.get('difficulty'),
                'points':           lab.get('total_points'),
                'register_mapping': reg_map,
                'initial_values':   init_vals,
                'test_cases':       test_cases,
            }

        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to load labs')}), 500


@app.route('/api/labs', methods=['POST'])
def create_lab():
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    data = request.get_json()
    if not data or 'lab_id' not in data or 'title' not in data:
        return jsonify({'error': 'Missing required fields'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        default_course_id = data.get('course_id')
        if not default_course_id:
            cursor.execute("SELECT course_id FROM courses LIMIT 1")
            course = cursor.fetchone()
            if course:
                default_course_id = course['course_id']
            else:
                cursor.execute("""
                    INSERT INTO courses (course_code, course_name, semester, year, is_active)
                    VALUES ('DEFAULT', 'Default Course', 'Fall', 2024, TRUE)
                    RETURNING course_id
                """)
                default_course_id = cursor.fetchone()['course_id']

        cursor.execute("""
            INSERT INTO labs (
                lab_id, course_id, title, description, instructions,
                starter_code, solution_code, register_mapping, initial_values,
                difficulty, total_points, max_instructions, time_limit_seconds,
                max_memory_kb, release_date, due_date, is_published,
                use_isolation
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING lab_id
        """, (
            data['lab_id'], default_course_id, data['title'],
            data.get('description'), data.get('instructions', ''),
            data.get('starter_code'), data.get('solution_code'),
            json.dumps(data.get('register_mapping')) if data.get('register_mapping') else None,
            json.dumps(data.get('initial_values')) if data.get('initial_values') else None,
            data.get('difficulty', 'beginner'), data.get('points', 100),
            data.get('max_instructions', 10000), data.get('time_limit_seconds', 10),
            data.get('max_memory_kb', 1024), data.get('release_date'),
            data.get('due_date'), data.get('is_published', True),
            data.get('use_isolation', False),
        ))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Lab created successfully', 'lab_id': data['lab_id']}), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        if 'IntegrityError' in type(e).__name__:
            return jsonify({'error': 'Lab ID already exists'}), 409
        return jsonify({'error': _safe_error(e, 'Failed to create lab')}), 500


@app.route('/api/labs/<lab_id>', methods=['PUT'])
def update_lab(lab_id):
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    data = request.get_json()
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            UPDATE labs SET
                title = COALESCE(%s, title),
                description = COALESCE(%s, description),
                instructions = COALESCE(%s, instructions),
                starter_code = COALESCE(%s, starter_code),
                solution_code = COALESCE(%s, solution_code),
                register_mapping = COALESCE(%s, register_mapping),
                initial_values = COALESCE(%s, initial_values),
                difficulty = COALESCE(%s, difficulty),
                total_points = COALESCE(%s, total_points),
                due_date = COALESCE(%s, due_date),
                is_published = COALESCE(%s, is_published),
                use_isolation = COALESCE(%s, use_isolation)
            WHERE lab_id = %s RETURNING lab_id
        """, (
            data.get('title'), data.get('description'), data.get('instructions'),
            data.get('starter_code'), data.get('solution_code'),
            json.dumps(data.get('register_mapping')) if 'register_mapping' in data else None,
            json.dumps(data.get('initial_values')) if 'initial_values' in data else None,
            data.get('difficulty'), data.get('points'), data.get('due_date'),
            data.get('is_published'), data.get('use_isolation'), lab_id,
        ))

        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Lab not found'}), 404

        conn.commit()
        conn.close()
        return jsonify({'message': 'Lab updated successfully'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to update lab')}), 500


@app.route('/api/labs/<lab_id>', methods=['DELETE'])
def delete_lab(lab_id):
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM submissions WHERE lab_id = %s", (lab_id,))
        cursor.execute("DELETE FROM lab_test_cases WHERE lab_id = %s", (lab_id,))
        cursor.execute("DELETE FROM labs WHERE lab_id = %s RETURNING lab_id", (lab_id,))

        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Lab not found'}), 404

        conn.commit()
        conn.close()
        return jsonify({'message': 'Lab deleted successfully'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to delete lab')}), 500


@app.route('/api/labs/<lab_id>/test-cases', methods=['GET'])
def get_lab_test_cases(lab_id):
    role = session.get('role', '')
    if role not in ('instructor', 'ta'):
        return jsonify({'error': 'Unauthorized'}), 403

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        test_cases = _fetch_test_cases(cursor, lab_id)
        conn.close()
        return jsonify({'lab_id': lab_id, 'test_cases': test_cases})
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to load test cases')}), 500


@app.route('/api/labs/<lab_id>/test-cases', methods=['POST'])
def create_test_case(lab_id):
    role = session.get('role', '')
    if role not in ('instructor', 'ta'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    if not data or 'test_name' not in data:
        return jsonify({'error': 'Missing required field: test_name'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT lab_id FROM labs WHERE lab_id = %s", (lab_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Lab not found'}), 404

        cursor.execute("""
            INSERT INTO lab_test_cases
                (lab_id, test_name, test_type, description,
                 input_data, expected_output, points, is_hidden, timeout_seconds)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING test_case_id
        """, (
            lab_id, data['test_name'], data.get('test_type', 'register'),
            data.get('description'),
            json.dumps(data.get('input_data', {})),
            json.dumps(data.get('expected_output', {})),
            data.get('points', 10), data.get('is_hidden', False),
            data.get('timeout_seconds', 5),
        ))
        new_id = str(cursor.fetchone()['test_case_id'])
        conn.commit()
        conn.close()
        return jsonify({'message': 'Test case created', 'test_case_id': new_id}), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to create test case')}), 500


@app.route('/api/labs/<lab_id>/test-cases/<test_case_id>', methods=['PUT'])
def update_test_case(lab_id, test_case_id):
    role = session.get('role', '')
    if role not in ('instructor', 'ta'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            UPDATE lab_test_cases SET
                test_name       = COALESCE(%s, test_name),
                test_type       = COALESCE(%s, test_type),
                description     = COALESCE(%s, description),
                input_data      = COALESCE(%s, input_data),
                expected_output = COALESCE(%s, expected_output),
                points          = COALESCE(%s, points),
                is_hidden       = COALESCE(%s, is_hidden),
                timeout_seconds = COALESCE(%s, timeout_seconds)
            WHERE test_case_id = %s AND lab_id = %s
            RETURNING test_case_id
        """, (
            data.get('test_name'), data.get('test_type'), data.get('description'),
            json.dumps(data['input_data']) if 'input_data' in data else None,
            json.dumps(data['expected_output']) if 'expected_output' in data else None,
            data.get('points'), data.get('is_hidden'), data.get('timeout_seconds'),
            test_case_id, lab_id,
        ))
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Test case not found'}), 404
        conn.commit()
        conn.close()
        return jsonify({'message': 'Test case updated', 'test_case_id': test_case_id})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to update test case')}), 500


@app.route('/api/labs/<lab_id>/test-cases/<test_case_id>', methods=['DELETE'])
def delete_test_case(lab_id, test_case_id):
    role = session.get('role', '')
    if role not in ('instructor', 'ta'):
        return jsonify({'error': 'Unauthorized'}), 403

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM lab_test_cases WHERE test_case_id = %s AND lab_id = %s RETURNING test_case_id",
            (test_case_id, lab_id))
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Test case not found'}), 404
        conn.commit()
        conn.close()
        return jsonify({'message': 'Test case deleted'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to delete test case')}), 500


@app.route('/api/students', methods=['GET'])
def list_students():
    role = session.get('role', '')
    if role not in ('instructor', 'ta'):
        return jsonify({'error': 'Unauthorized'}), 403

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT u.user_id, u.username, u.full_name, u.email, u.asu_id,
                   u.created_at, u.last_login,
                   COUNT(DISTINCT s.lab_id)    AS labs_attempted,
                   COUNT(s.submission_id)       AS total_submissions,
                   COALESCE(ROUND(AVG(
                     CASE WHEN s.total_possible > 0
                          THEN s.score * 100.0 / s.total_possible ELSE NULL END
                   ), 1), 0)                    AS avg_score_pct,
                   MAX(s.submitted_at)          AS last_submission
            FROM users u
            LEFT JOIN submissions s ON s.user_id = u.user_id
            WHERE u.role = 'student' AND u.is_active = true
            GROUP BY u.user_id
            ORDER BY u.full_name, u.username
        """)
        students = cursor.fetchall()

        cursor.execute("SELECT COUNT(*) AS cnt FROM labs WHERE is_published = true")
        total_labs = cursor.fetchone()['cnt']

        conn.close()

        return jsonify([{
            'user_id':           str(s['user_id']),
            'username':          s['username'],
            'full_name':         s['full_name'],
            'email':             s['email'],
            'asu_id':            s['asu_id'],
            'created_at':        s['created_at'].isoformat() if s['created_at'] else None,
            'last_login':        s['last_login'].isoformat() if s['last_login'] else None,
            'labs_attempted':    s['labs_attempted'],
            'total_labs':        total_labs,
            'total_submissions': s['total_submissions'],
            'avg_score_pct':     float(s['avg_score_pct']),
            'last_submission':   s['last_submission'].isoformat() if s['last_submission'] else None,
        } for s in students])
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to load students')}), 500


@app.route('/api/students/<user_id>', methods=['GET'])
def get_student_detail(user_id):
    role = session.get('role', '')
    if role not in ('instructor', 'ta'):
        return jsonify({'error': 'Unauthorized'}), 403

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT user_id, username, full_name, email, asu_id, created_at, last_login
            FROM users WHERE user_id = %s
        """, (user_id,))
        student = cursor.fetchone()
        if not student:
            conn.close()
            return jsonify({'error': 'Student not found'}), 404

        cursor.execute("""
            SELECT lab_id, title, total_points, difficulty
            FROM labs WHERE is_published = true ORDER BY lab_id
        """)
        all_labs = cursor.fetchall()

        cursor.execute("""
            SELECT submission_id, lab_id, score, total_possible,
                   test_results, submitted_at, source_code,
                   duration_seconds, run_count, started_at, timing_flagged
            FROM submissions WHERE user_id = %s ORDER BY submitted_at DESC
        """, (user_id,))
        submissions = cursor.fetchall()

        # --- NEW TELEMETRY QUERY ---
        cursor.execute("""
            SELECT lab_id, source_code, executed_at, is_step
            FROM run_telemetry 
            WHERE user_id = %s 
            ORDER BY executed_at DESC
        """, (user_id,))
        telemetry = cursor.fetchall()
        conn.close()

        subs_by_lab = {}
        for sub in submissions:
            lid = sub['lab_id']
            subs_by_lab.setdefault(lid, []).append({
                'submission_id':    str(sub['submission_id']),
                'score':            sub['score'],
                'total_possible':   sub['total_possible'],
                'test_results':     sub['test_results'],
                'submitted_at':     sub['submitted_at'].isoformat() if sub['submitted_at'] else None,
                'source_code':      sub['source_code'],
                'duration_seconds': sub.get('duration_seconds'),
                'run_count':        sub.get('run_count'),
                'started_at':       sub['started_at'].isoformat() if sub.get('started_at') else None,
                'timing_flagged':   sub.get('timing_flagged', False),
            })

        tel_by_lab = {}
        for t in telemetry:
            tel_by_lab.setdefault(t['lab_id'], []).append({
                'source_code': t['source_code'],
                'executed_at': t['executed_at'].isoformat(),
                'is_step': t['is_step']
            })

        lab_details = []
        for lab in all_labs:
            lid = lab['lab_id']
            lab_subs = subs_by_lab.get(lid, [])
            lab_tel  = tel_by_lab.get(lid, []) # NEW
            best_score = max((s['score'] for s in lab_subs), default=None)
            best_possible = lab_subs[0]['total_possible'] if lab_subs else lab['total_points']

            lab_details.append({
                'lab_id': lid, 'title': lab['title'],
                'difficulty': lab['difficulty'], 'total_points': lab['total_points'],
                'attempt_count': len(lab_subs), 'best_score': best_score,
                'best_possible': best_possible,
                'latest_submission': lab_subs[0] if lab_subs else None,
                'submissions': lab_subs,
                'telemetry': lab_tel
            })

        return jsonify({
            'student': {
                'user_id':    str(student['user_id']),
                'username':   student['username'],
                'full_name':  student['full_name'],
                'email':      student['email'],
                'asu_id':     student['asu_id'],
                'created_at': student['created_at'].isoformat() if student['created_at'] else None,
                'last_login': student['last_login'].isoformat() if student['last_login'] else None,
            },
            'labs': lab_details,
        })
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to load student details')}), 500


@app.route('/api/roster', methods=['GET'])
def get_roster():
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT r.roster_id, r.asurite, r.asu_id, r.full_name, r.email,
                   r.is_registered, r.added_at,
                   u.user_id AS registered_user_id
            FROM course_roster r
            LEFT JOIN users u ON u.username = r.asurite AND u.role = 'student'
            ORDER BY r.full_name, r.asurite
        """)
        rows = cursor.fetchall()
        conn.close()

        return jsonify([{
            'roster_id':    str(r['roster_id']),
            'asurite':      r['asurite'],
            'asu_id':       r['asu_id'],
            'full_name':    r['full_name'],
            'email':        r['email'],
            'is_registered': r['is_registered'] or r['registered_user_id'] is not None,
            'added_at':     r['added_at'].isoformat() if r['added_at'] else None,
        } for r in rows])
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to load roster')}), 500


@app.route('/api/roster/upload', methods=['POST'])
def upload_roster():
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    filename = (file.filename or '').lower()

    try:
        rows = []
        if filename.endswith('.csv'):
            text = file.read().decode('utf-8-sig')
            reader = csv.DictReader(StringIO(text))
            for r in reader:
                rows.append({k.strip().lower(): (v or '').strip() for k, v in r.items()})
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            if not EXCEL_AVAILABLE:
                return jsonify({'error': 'Excel support not available on server'}), 500
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h or '').strip().lower() for h in header_row]
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        d[headers[i]] = str(val or '').strip()
                rows.append(d)
            wb.close()
        else:
            return jsonify({'error': 'Unsupported file type. Use .csv or .xlsx'}), 400

        if not rows:
            return jsonify({'error': 'File is empty'}), 400

        COL_MAP_ASURITE = {'asurite', 'asurite id', 'asuriteid', 'username'}
        COL_MAP_ASUID   = {'asu id', 'asuid', 'asu_id', 'student id', 'id number', 'id'}
        COL_MAP_NAME    = {'full name', 'fullname', 'full_name', 'name', 'student name', 'student'}
        COL_MAP_EMAIL   = {'email', 'asu email', 'e-mail'}

        def find_col(row_keys, candidates):
            for k in row_keys:
                if k in candidates:
                    return k
            return None

        sample_keys = list(rows[0].keys())
        col_asurite = find_col(sample_keys, COL_MAP_ASURITE)
        col_asuid   = find_col(sample_keys, COL_MAP_ASUID)
        col_name    = find_col(sample_keys, COL_MAP_NAME)
        col_email   = find_col(sample_keys, COL_MAP_EMAIL)

        if not col_asurite:
            return jsonify({'error': 'Could not find an ASURITE column. Expected one of: ' + ', '.join(sorted(COL_MAP_ASURITE))}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Database connection failed'}), 500

        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT course_id FROM courses LIMIT 1")
        course = cursor.fetchone()
        if not course:
            conn.close()
            return jsonify({'error': 'No course exists. Create a course first.'}), 400
        course_id = course['course_id']

        added = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows):
            asurite = row.get(col_asurite, '').strip().lower()
            if not asurite:
                skipped += 1
                continue

            asu_id    = row.get(col_asuid, '').strip() if col_asuid else ''
            full_name = row.get(col_name, '').strip() if col_name else ''
            email     = row.get(col_email, '').strip().lower() if col_email else ''
            if not email and asurite:
                email = asurite + '@asu.edu'

            try:
                cursor.execute("""
                    INSERT INTO course_roster (course_id, asurite, asu_id, full_name, email, added_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (asurite, course_id) DO UPDATE SET
                        asu_id = EXCLUDED.asu_id,
                        full_name = EXCLUDED.full_name,
                        email = EXCLUDED.email
                """, (course_id, asurite, asu_id, full_name, email, uid))
                added += 1
            except Exception as row_err:
                errors.append(f"Row {i+2}: {row_err}")
                conn.rollback()

        conn.commit()
        conn.close()

        return jsonify({
            'message': f'Roster uploaded: {added} added/updated, {skipped} skipped',
            'added': added,
            'skipped': skipped,
            'errors': errors[:10],
        })
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to process roster file')}), 500


@app.route('/api/roster/<roster_id>', methods=['DELETE'])
def delete_roster_entry(roster_id):
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM course_roster WHERE roster_id = %s RETURNING roster_id", (roster_id,))
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Entry not found'}), 404
        conn.commit()
        conn.close()
        return jsonify({'message': 'Roster entry deleted'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to delete roster entry')}), 500


@app.route('/api/roster/clear', methods=['DELETE'])
def clear_roster():
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM course_roster")
        count = cursor.rowcount
        conn.commit()
        conn.close()
        return jsonify({'message': f'Cleared {count} roster entries'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': _safe_error(e, 'Failed to clear roster')}), 500


@app.route('/api/roster/template', methods=['GET'])
def roster_template():
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 403 if err == 'not_teacher' else 401

    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel support not available on server'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    headers = ['ASURITE', 'ASU ID', 'Full Name', 'Email']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    example = ['jdoe42', '1234567890', 'Jane Doe', 'jdoe42@asu.edu']
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=val)

    for col_idx, width in enumerate([15, 15, 25, 30], 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='roster_template.xlsx'
    )


@app.route('/api/export/grades/<lab_id>', methods=['GET'])
def export_grades_excel(lab_id):
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 401

    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        if lab_id == "ALL":
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id, s.lab_id)
                    u.username, u.user_id, u.email, s.lab_id, s.score,
                    s.total_possible AS max_score,
                    (s.score::float / NULLIF(s.total_possible, 0) * 100) AS percentage,
                    s.submitted_at
                FROM submissions s JOIN users u ON s.user_id = u.user_id
                ORDER BY s.user_id, s.lab_id, s.score DESC, s.submitted_at DESC
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id)
                    u.username, u.user_id, u.email, s.score,
                    s.total_possible AS max_score,
                    (s.score::float / NULLIF(s.total_possible, 0) * 100) AS percentage,
                    s.submitted_at
                FROM submissions s JOIN users u ON s.user_id = u.user_id
                WHERE s.lab_id = %s
                ORDER BY s.user_id, s.score DESC, s.submitted_at DESC
            """, (lab_id,))

        submissions = cursor.fetchall()
        conn.close()

        if not submissions:
            return jsonify({'error': 'No submissions found'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"

        if lab_id == "ALL":
            headers = ['Student Name', 'Student ID', 'Email', 'Lab', 'Auto Score',
                       'Max Score', 'Auto %', 'Manual Adj', 'Final Score', 'Final %', 'Submitted']
        else:
            headers = ['Student Name', 'Student ID', 'Email', 'Auto Score',
                       'Max Score', 'Auto %', 'Manual Adj', 'Final Score', 'Final %', 'Submitted']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, sub in enumerate(submissions, 2):
            col = 1
            ws.cell(row=row_idx, column=col, value=sub.get('username', 'Unknown')); col += 1
            ws.cell(row=row_idx, column=col, value=str(sub['user_id'])); col += 1
            ws.cell(row=row_idx, column=col, value=sub.get('email', '')); col += 1

            if lab_id == "ALL":
                ws.cell(row=row_idx, column=col, value=sub.get('lab_id', '')); col += 1

            auto_score_col = col
            ws.cell(row=row_idx, column=col, value=sub['score']); col += 1
            max_score_col = col
            ws.cell(row=row_idx, column=col, value=sub['max_score']); col += 1

            pct = float(sub.get('percentage') or 0)
            pct_cell = ws.cell(row=row_idx, column=col, value=pct)
            if pct >= 90:
                pct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif pct >= 70:
                pct_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                pct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            col += 1

            adj_cell = ws.cell(row=row_idx, column=col, value=0.0)
            adj_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            adj_cell.font = Font(bold=True)
            adj_col = col; col += 1

            final_cell = ws.cell(row=row_idx, column=col)
            final_cell.value = f"={chr(64+auto_score_col)}{row_idx}+{chr(64+adj_col)}{row_idx}"
            final_cell.font = Font(bold=True)
            final_col = col; col += 1

            final_pct_cell = ws.cell(row=row_idx, column=col)
            final_pct_cell.value = f"=({chr(64+final_col)}{row_idx}/{chr(64+max_score_col)}{row_idx})*100"
            final_pct_cell.number_format = '0.0'
            final_pct_cell.font = Font(bold=True)
            col += 1

            ws.cell(row=row_idx, column=col, value=sub['submitted_at'].strftime('%Y-%m-%d %H:%M'))

        for column in ws.columns:
            col_letter = column[0].column_letter
            max_len = max((len(str(cell.value or '')) for cell in column), default=0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'all_grades_{ts}.xlsx' if lab_id == "ALL" else f'grades_{lab_id}_{ts}.xlsx'
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=fname)
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to export grades')}), 500


@app.route('/api/export/submissions-zip/<lab_id>', methods=['GET'])
def export_submissions_zip(lab_id):
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        if lab_id == "ALL":
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id, s.lab_id)
                    u.username, u.user_id, s.lab_id,
                    s.source_code, s.score, s.submitted_at
                FROM submissions s JOIN users u ON s.user_id = u.user_id
                ORDER BY s.user_id, s.lab_id, s.score DESC
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT ON (s.user_id)
                    u.username, u.user_id, s.lab_id,
                    s.source_code, s.score, s.submitted_at
                FROM submissions s JOIN users u ON s.user_id = u.user_id
                WHERE s.lab_id = %s ORDER BY s.user_id, s.score DESC
            """, (lab_id,))

        submissions = cursor.fetchall()
        conn.close()

        if not submissions:
            return jsonify({'error': 'No submissions found'}), 404

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for sub in submissions:
                fname = f"{sub['username']}_{sub['lab_id']}_score{sub['score']}.asm"
                code = (sub['source_code'] or '').replace('\\n', '\n')
                content = (
                    f"# Student: {sub['username']}\n"
                    f"# Lab: {sub['lab_id']}\n"
                    f"# Score: {sub['score']}\n"
                    f"# Submitted: {sub['submitted_at']}\n"
                    f"# ==========================================\n\n"
                    f"{code}\n"
                )
                zf.writestr(fname, content)

        zip_buffer.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'all_submissions_{ts}.zip' if lab_id == "ALL" else f'submissions_{lab_id}_{ts}.zip'
        return send_file(zip_buffer, mimetype='application/zip',
                         as_attachment=True, download_name=fname)
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to create ZIP')}), 500


@app.route('/api/export/submissions/<lab_id>', methods=['GET'])
def export_submissions_csv(lab_id):
    uid, err = _require_teacher()
    if err:
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        if lab_id == "ALL":
            cursor.execute("""
                SELECT u.username, u.user_id, u.email,
                       s.lab_id, s.score, s.total_possible,
                       s.source_code, s.submitted_at
                FROM submissions s JOIN users u ON s.user_id = u.user_id
                ORDER BY u.username, s.lab_id, s.submitted_at DESC
            """)
        else:
            cursor.execute("""
                SELECT u.username, u.user_id, u.email,
                       s.score, s.total_possible,
                       s.source_code, s.submitted_at
                FROM submissions s JOIN users u ON s.user_id = u.user_id
                WHERE s.lab_id = %s
                ORDER BY u.username, s.submitted_at DESC
            """, (lab_id,))

        submissions = cursor.fetchall()
        conn.close()

        if not submissions:
            return jsonify({'error': 'No submissions found'}), 404

        output = StringIO()
        writer = csv.writer(output)

        if lab_id == "ALL":
            writer.writerow(['Student Name', 'Student ID', 'Email', 'Lab', 'Score',
                             'Max Score', 'Percentage', 'Submitted At', 'Source Code'])
        else:
            writer.writerow(['Student Name', 'Student ID', 'Email', 'Score',
                             'Max Score', 'Percentage', 'Submitted At', 'Source Code'])

        for sub in submissions:
            pct = (sub['score'] / sub['total_possible'] * 100) if sub['total_possible'] > 0 else 0
            row = [sub.get('username', 'Unknown'), str(sub['user_id']), sub.get('email', '')]
            if lab_id == "ALL":
                row.append(sub.get('lab_id', ''))
            row.extend([
                sub['score'], sub['total_possible'], f"{pct:.1f}%",
                sub['submitted_at'].strftime('%Y-%m-%d %H:%M:%S'),
                sub['source_code'].replace('\\n', '\n').replace('\n', ' | ')
            ])
            writer.writerow(row)

        output.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'all_history_{ts}.csv' if lab_id == "ALL" else f'history_{lab_id}_{ts}.csv'
        return send_file(BytesIO(output.getvalue().encode('utf-8')),
                         mimetype='text/csv', as_attachment=True, download_name=fname)
    except Exception as e:
        return jsonify({'error': _safe_error(e, 'Failed to export CSV')}), 500